from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
from datetime import datetime


def generate_excel(report):
    folder = "app/reports/generated"

    os.makedirs(folder, exist_ok=True)

    file_path = os.path.join(
        folder,
        "textile_waste_report.xlsx"
    )

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Textile Waste Report"

    title = "Textile Waste Intelligence Report"

    sheet.merge_cells("A1:B1")

    sheet["A1"] = title
    sheet["A1"].font = Font(
        bold=True,
        size=18
    )
    sheet["A1"].alignment = Alignment(
        horizontal="center"
    )

    sheet.merge_cells("A2:B2")

    sheet["A2"] = (
        "Generated On: "
        + datetime.now().strftime(
            "%d-%m-%Y %H:%M"
        )
    )

    sheet["A2"].alignment = Alignment(
        horizontal="center"
    )

    current_row = 4

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    for section, data in report.items():

        sheet.merge_cells(
            start_row=current_row,
            start_column=1,
            end_row=current_row,
            end_column=2
        )

        section_cell = sheet.cell(
            row=current_row,
            column=1
        )

        section_cell.value = section

        section_cell.font = Font(
            bold=True,
            size=13
        )

        section_cell.fill = PatternFill(
            "solid",
            fgColor="0F766E"
        )

        section_cell.font = Font(
            bold=True,
            color="FFFFFF",
            size=13
        )

        section_cell.alignment = Alignment(
            horizontal="left"
        )

        current_row += 1

        parameter_cell = sheet.cell(
            row=current_row,
            column=1
        )

        result_cell = sheet.cell(
            row=current_row,
            column=2
        )

        parameter_cell.value = "Parameter"
        result_cell.value = "Result"

        parameter_cell.font = Font(
            bold=True
        )

        result_cell.font = Font(
            bold=True
        )

        parameter_cell.fill = PatternFill(
            "solid",
            fgColor="D9EAD3"
        )

        result_cell.fill = PatternFill(
            "solid",
            fgColor="D9EAD3"
        )

        parameter_cell.border = thin_border
        result_cell.border = thin_border

        current_row += 1

        for key, value in data.items():

            parameter_cell = sheet.cell(
                row=current_row,
                column=1
            )

            result_cell = sheet.cell(
                row=current_row,
                column=2
            )

            parameter_cell.value = key
            result_cell.value = str(value)

            parameter_cell.border = thin_border
            result_cell.border = thin_border

            parameter_cell.alignment = Alignment(
                vertical="top"
            )

            result_cell.alignment = Alignment(
                vertical="top",
                wrap_text=True
            )

            current_row += 1

        current_row += 1

    sheet.column_dimensions["A"].width = 35
    sheet.column_dimensions["B"].width = 65

    sheet.freeze_panes = "A4"

    workbook.save(file_path)

    return file_path